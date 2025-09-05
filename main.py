import os, io, time, json, logging, re, uuid, csv, tempfile, gc, asyncio
from typing import List, Tuple, Optional, Iterable, Dict, Any, Callable
from dataclasses import dataclass, field
from enum import Enum

import httpx
import pandas as pd
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse, RedirectResponse, JSONResponse, PlainTextResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

from azure.storage.blob import BlobServiceClient, ContentSettings
from azure.identity import DefaultAzureCredential
from azure.core.exceptions import ResourceExistsError

# Excel streaming
from openpyxl import load_workbook

# offload heavy sync steps to a thread so the event loop stays alive
import anyio

# ============================= In-memory log buffer =============================
class RingLogHandler(logging.Handler):
    def __init__(self, capacity: int = 2000):
        super().__init__()
        self.capacity = max(100, int(capacity))
        self.buf: List[Dict[str, Any]] = []
        self._lock = asyncio.Lock()

    def format_record(self, record: logging.LogRecord) -> Dict[str, Any]:
        try:
            msg = record.getMessage()
        except Exception:
            msg = record.msg
        parsed = None
        if isinstance(record.args, dict):
            parsed = record.args
        else:
            try:
                parsed = json.loads(msg) if isinstance(msg, str) and msg.startswith("{") else None
            except Exception:
                parsed = None
        return {
            "ts": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime(record.created)),
            "level": record.levelname,
            "logger": record.name,
            "message": msg if parsed is None else parsed,
        }

    async def aemit(self, record: logging.LogRecord):
        item = self.format_record(record)
        async with self._lock:
            self.buf.append(item)
            if len(self.buf) > self.capacity:
                self.buf = self.buf[-self.capacity:]

    def emit(self, record: logging.LogRecord):
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                asyncio.create_task(self.aemit(record))
            else:
                item = self.format_record(record)
                self.buf.append(item)
                if len(self.buf) > self.capacity:
                    self.buf = self.buf[-self.capacity:]
        except RuntimeError:
            item = self.format_record(record)
            self.buf.append(item)
            if len(self.buf) > self.capacity:
                self.buf = self.buf[-self.capacity:]


ring_handler = RingLogHandler(capacity=3000)
logging.basicConfig(level=logging.INFO, handlers=[ring_handler])
log = logging.getLogger("fiquebot")
log.setLevel(logging.INFO)
if ring_handler not in log.handlers:
    log.addHandler(ring_handler)
log.info("fiquebot backend starting with in-memory log buffer (capacity=%d)", ring_handler.capacity)

# ============================= App init & CORS =============================
app = FastAPI(title="Fiquebot API", version="2.3.0")

_default_origins = [
    "http://localhost:8000",
    "http://127.0.0.1:8000",
    "http://localhost:5500",
    "http://127.0.0.1:5500",
    "https://<YOUR-STATIC-WEB-APP>.azurestaticapps.net",
    "https://orange-dune-0ea42a603.1.azurestaticapps.net",
]
_env_origins = os.environ.get("FRONTEND_ORIGINS") or os.environ.get("ALLOWED_ORIGINS")
allow_origins = [o.strip() for o in _env_origins.split(",") if o.strip()] if _env_origins else _default_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve /ui (optional)
if os.path.isdir("web"):
    app.mount("/ui", StaticFiles(directory="web", html=True), name="ui")

    @app.get("/", include_in_schema=False)
    async def _root():
        return RedirectResponse(url="/ui/")
else:
    @app.get("/", include_in_schema=False)
    async def _root_missing():
        return JSONResponse(
            {"message": "Frontend not found. Create web/index.html or deploy Azure Static Web App."},
            status_code=200,
        )

# ============================= Config =============================
TRN_EP = os.environ.get("AZURE_TRANSLATOR_ENDPOINT", "https://api.cognitive.microsofttranslator.com").rstrip("/")
TRN_KEY = os.environ.get("AZURE_TRANSLATOR_KEY", "")
TRN_REGION = os.environ.get("AZURE_TRANSLATOR_REGION", "westeurope")

# === UPDATED DEFAULTS FOR 4.1 MINI + STRUCTURED OUTPUTS ===
AOAI_EP = (os.environ.get("AZURE_OPENAI_ENDPOINT", "") or "").rstrip("/")
AOAI_DEP = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4.1-mini")
AOAI_VER = os.environ.get("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
AOAI_KEY = os.environ.get("AZURE_OPENAI_API_KEY", "")

STORAGE_ACCOUNT = os.environ.get("STORAGE_ACCOUNT_NAME", "fiqueuploadstore")

TRANSLATE_PROVIDER = os.environ.get("TRANSLATE_PROVIDER", "llm").lower()  # "llm" | "ms"
LLM_MAX_BATCH = int(os.environ.get("LLM_MAX_BATCH", "80"))  # balance speed/peak memory
LLM_SYS = "You output ONLY valid JSON that matches the provided schema. No commentary."
ENTITY_CONF_THRESHOLD = float(os.environ.get("ENTITY_CONF_THRESHOLD", "0.60"))
CSV_CHUNK = int(os.environ.get("CSV_CHUNK", "10000"))
LLM_SEED = int(os.environ.get("LLM_SEED", "7"))

# ============================= Small utils =============================
def _soft_require(cond: bool, msg: str) -> bool:
    if not cond:
        log.warning({"event":"soft-require-failed","msg": msg})
        return False
    return True

def is_truthy(s: Optional[str]) -> bool:
    return str(s or "").strip().lower() in {"1", "true", "yes", "y"}

def json_log(event: str, **fields):
    rec = {"event": event, **fields}
    log.info(json.dumps(rec))

# ============================= Dialing codes (condensed) =============================
_NANP = {
    "united states", "united states of america", "usa", "canada",
    "bahamas", "barbados", "bermuda", "jamaica", "dominican republic", "haiti",
    "trinidad and tobago", "puerto rico", "grenada", "saint lucia",
    "antigua and barbuda", "saint kitts and nevis", "saint vincent and the grenadines",
    "anguilla", "cayman islands", "turks and caicos islands", "dominica",
    "british virgin islands", "us virgin islands", "guam", "northern mariana islands"
}
_COUNTRY_TO_DIAL = {
    "india": "+91", "kenya": "+254", "nepal": "+977", "bangladesh": "+880", "pakistan": "+92",
    "sri lanka": "+94", "australia": "+61", "new zealand": "+64", "united kingdom": "+44", "uk": "+44",
    "england": "+44", "ireland": "+353", "south africa": "+27", "nigeria": "+234", "ghana": "+233",
    "tanzania": "+255", "uganda": "+256", "rwanda": "+250", "ethiopia": "+251", "zambia": "+260",
    "zimbabwe": "+263", "botswana": "+267", "namibia": "+264", "morocco": "+212", "tunisia": "+216",
    "algeria": "+213", "egypt": "+20", "saudi arabia": "+966", "united arab emirates": "+971", "uae": "+971",
    "qatar": "+974", "oman": "+968", "bahrain": "+973", "iran": "+98", "iraq": "+964", "turkey": "+90",
    "israel": "+972", "jordan": "+962", "lebanon": "+961",
    "china": "+86", "japan": "+81", "south korea": "+82", "korea, republic of": "+82", "north korea": "+850",
    "hong kong": "+852", "macau": "+853", "taiwan": "+886", "vietnam": "+84", "thailand": "+66",
    "laos": "+856", "cambodia": "+855", "malaysia": "+60", "singapore": "+65", "indonesia": "+62",
    "philippines": "+63", "myanmar": "+95", "brunei": "+673", "mongolia": "+976", "afghanistan": "+93",
    "france": "+33", "germany": "+49", "italy": "+39", "spain": "+34", "portugal": "+351",
    "netherlands": "+31", "belgium": "+32", "luxembourg": "+352", "switzerland": "+41", "austria": "+43",
    "poland": "+48", "czech republic": "+420", "czechia": "+420", "slovakia": "+421", "hungary": "+36",
    "romania": "+40", "bulgaria": "+359", "greece": "+30", "croatia": "+385", "slovenia": "+386",
    "serbia": "+381", "bosnia and herzegovina": "+387", "north macedonia": "+389", "albania": "+355",
    "iceland": "+354", "norway": "+47", "sweden": "+46", "finland": "+358", "denmark": "+45",
    "estonia": "+372", "latvia": "+371", "lithuania": "+370", "ukraine": "+380", "belarus": "+375",
    "moldova": "+373", "georgia": "+995", "armenia": "+374", "azerbaijan": "+994", "russia": "+7", "kazakhstan": "+7",
    "mexico": "+52", "guatemala": "+502", "belize": "+501", "honduras": "+504", "el salvador": "+503",
    "nicaragua": "+505", "costa rica": "+506", "panama": "+507", "colombia": "+57", "venezuela": "+58",
    "ecuador": "+593", "peru": "+51", "bolivia": "+591", "paraguay": "+595", "chile": "+56",
    "argentina": "+54", "uruguay": "+598", "brazil": "+55", "cuba": "+53",
    "dominican republic": "+1", "jamaica": "+1", "trinidad and tobago": "+1", "barbados": "+1", "bahamas": "+1",
    "canada": "+1", "united states": "+1", "united states of america": "+1", "usa": "+1",
}
_DIAL_PREFIXES = sorted({v.replace("+", "").replace("-", "") for v in _COUNTRY_TO_DIAL.values()},
                        key=lambda x: (-len(x), x))

def _norm_country(name: str) -> str:
    s = (name or "").strip().lower()
    s = s.replace("&", "and")
    s = re.sub(r"\s+", " ", s)
    return s

def _infer_from_phone(phone: str) -> str:
    if not phone:
        return ""
    digits = re.sub(r"[^\d]", "", phone)
    if phone.strip().startswith("00"):
        digits = digits[2:]
    for pref in _DIAL_PREFIXES:
        if digits.startswith(pref):
            return pref
    return ""

def country_to_dial(country: str, phone: str = "") -> str:
    if not country and phone:
        code = _infer_from_phone(phone)
        return f"+{code}" if code else ""
    s = _norm_country(country)
    if s in _NANP or any(k in s for k in ["usa", "united states"]):
        return "+1"
    if s in _COUNTRY_TO_DIAL:
        return _COUNTRY_TO_DIAL[s]
    if phone:
        code = _infer_from_phone(phone)
        return f"+{code}" if code else ""
    return ""

# ============================= Blob helpers =============================
def get_blob_client() -> Optional[BlobServiceClient]:
    if not _soft_require(STORAGE_ACCOUNT, "Storage account not configured"):
        return None
    connection_string = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
    try:
        if connection_string:
            log.info("Using connection string for Blob Service Client")
            return BlobServiceClient.from_connection_string(connection_string)
        else:
            log.info("Using DefaultAzureCredential for Blob Service Client")
            credential = DefaultAzureCredential()
            return BlobServiceClient(f"https://{STORAGE_ACCOUNT}.blob.core.windows.net", credential=credential)
    except Exception as e:
        log.warning({"event":"blob-client-fallback", "error": str(e)})
        return None

def ensure_container(blob_service: BlobServiceClient, name: str):
    try:
        blob_service.create_container(name)
        log.info({"event":"container-created", "name": name})
    except ResourceExistsError:
        pass
    except Exception as e:
        log.warning({"event":"container-create-skip", "name": name, "error": str(e)})

# ============================= HTTP helpers =============================
def httpx_client(timeout: int = 90) -> httpx.Client:
    return httpx.Client(timeout=timeout)

# ============================= Translator (MS) =============================
def _translate_and_detect_ms(texts: List[str], to_lang: str = "en") -> List[dict]:
    if not (TRN_EP and TRN_KEY and TRN_REGION):
        return [{"translated": t or "", "lang": "en", "confidence": 1.0} for t in texts]

    url = f"{TRN_EP}/translate?api-version=3.0&to={to_lang}"
    headers = {
        "Ocp-Apim-Subscription-Key": TRN_KEY,
        "Ocp-Apim-Subscription-Region": TRN_REGION,
        "Content-Type": "application/json",
    }
    out: List[dict] = []
    for i in range(0, len(texts), 50):
        batch = texts[i : i + 50]
        payload = [{"Text": t or ""} for t in batch]
        try:
            with httpx_client(90) as h:
                r = h.post(url, headers=headers, json=payload)
                r.raise_for_status()
                data = r.json()
            for item in data:
                t_en = item["translations"][0]["text"]
                det = item.get("detectedLanguage") or {}
                lang = (det.get("language", "") or "").lower()
                conf = float(det.get("score", det.get("confidence", 0)) or 0)
                out.append({"translated": t_en, "lang": lang, "confidence": conf})
        except Exception as e:
            log.warning({"event":"translate-fallback-ms-batch", "error": str(e)})
            out.extend([{"translated": t or "", "lang": "en", "confidence": 0.0} for t in batch])
    return out

# ============================= LLM batched (Azure OpenAI, 4.1 structured) =============================
def _rows_schema(strict: bool = True) -> Dict[str, Any]:
    """JSON Schema used for structured outputs"""
    return {
        "type": "object",
        "properties": {
            "rows": {
                "type": "array",
                "items": {
                    "type": "object",
                    "required": ["id","translated","lang","confidence","country","phone","book","language_mentioned","address"],
                    "properties": {
                        "id": {"type":"integer"},
                        "translated": {"type":"string"},
                        "lang": {"type":"string"},
                        "confidence": {"type":"number"},
                        "country": {"type":"string"},
                        "phone": {"type":"string"},
                        "book": {"type":"string", "enum":["Gyan Ganga","Way of Living",""]},
                        "language_mentioned": {"type":"string"},
                        "address": {"type":"string"}
                    },
                    "additionalProperties": False
                }
            }
        },
        "required": ["rows"],
        "additionalProperties": False
    }

def _llm_chat_jsonl(prompt: str, temperature: float = 0, max_tokens: int = 3300) -> List[Dict[str, Any]]:
    """
    With GPT-4.1+ structured outputs, this returns a list of row dicts from a single JSON object: {"rows":[...]}.
    """
    if not (AOAI_EP and AOAI_KEY and AOAI_DEP):
        raise RuntimeError("LLM not configured")

    url = f"{AOAI_EP}/openai/deployments/{AOAI_DEP}/chat/completions?api-version={AOAI_VER}"
    headers = {"Content-Type": "application/json", "api-key": AOAI_KEY}
    body = {
        "messages": [
            {"role": "system", "content": LLM_SYS},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0,        # deterministic
        "top_p": 0,              # deterministic
        "seed": LLM_SEED,        # reproducible
        "max_tokens": max_tokens,
        "response_format": {
            "type": "json_schema",
            "json_schema": {
                "name": "rows_schema",
                "schema": _rows_schema(strict=True),
                "strict": True
            }
        }
    }

    for attempt in range(3):
        try:
            with httpx_client(120) as h:
                r = h.post(url, headers=headers, json=body)
                r.raise_for_status()
                j = r.json()
            content = (j["choices"][0]["message"]["content"] or "").strip()
            obj = json.loads(content)
            rows = obj.get("rows", [])
            if not isinstance(rows, list):
                raise ValueError("structured output missing 'rows' array")
            return rows
        except Exception as e:
            wait = min(2 ** attempt, 8)
            log.warning({"event":"llm-retry","attempt":attempt+1,"wait":wait,"error":str(e)})
            time.sleep(wait)

    raise RuntimeError("LLM call failed after retries")

def _batch_rows_to_lines(rows: List[Dict[str, Any]]) -> str:
    def esc(s: str) -> str:
        s = (s or "").replace("\\", "\\\\").replace('"', '\\"')
        return s
    return "\n".join([f'- id="{r["id"]}" text="{esc(r["text"])}"' for r in rows])

def llm_translate_and_extract_batch(texts: List[str], to_lang: str = "en") -> List[dict]:
    """
    Single-pass: detect language, translate to en, extract entities. Uses structured outputs ("rows":[...]).
    """
    rows = [{"id": i, "text": (t or "")} for i, t in enumerate(texts)]
    lines = _batch_rows_to_lines(rows)
    prompt = f"""
You are a professional translator and information extractor.

For each input row:
1) Detect the language of "text".
2) Translate "text" to {to_lang}.
3) Extract entities from the translated English text:
   - country, phone, book ("Gyan Ganga"|"Way of Living"|""), language_mentioned, address.

Output a single JSON object with a top-level key "rows", where each item is:
{{"id": number, "translated": string, "lang": string, "confidence": number, "country": string, "phone": string, "book": string, "language_mentioned": string, "address": string}}

Rules:
- If unknown, use "" (empty string). Do NOT invent values.
- Never include any keys beyond the schema.
- The "lang" is the detected source language (ISO-ish code, lowercase).
- "confidence" ∈ [0,1].

Rows:
{lines}
""".strip()

    output_rows = _llm_chat_jsonl(prompt, temperature=0, max_tokens=3300)
    out_map = {int(o.get("id", -1)): o for o in output_rows if "id" in o}
    aligned: List[dict] = []
    for i, t in enumerate(texts):
        o = out_map.get(i) or {}
        aligned.append({
            "translated": str(o.get("translated", t or "")),
            "lang": str(o.get("lang", "")),
            "confidence": float(o.get("confidence", 0.0) or 0.0),
            "country": str(o.get("country", "")),
            "phone": str(o.get("phone", "")),
            "book": str(o.get("book", "")),
            "language_mentioned": str(o.get("language_mentioned", "")),
            "address": str(o.get("address", "")),
        })
    return aligned

def llm_entities_only_batch(texts_en: List[str]) -> List[dict]:
    """
    Entities-only: text is already English. Uses structured outputs with the same schema (translated/lang/confidence still present).
    """
    rows = [{"id": i, "text": (t or "")} for i, t in enumerate(texts_en)]
    lines = _batch_rows_to_lines(rows)
    prompt = f"""
You are an information extractor. Each row "text" is already English.

Extract: country, phone, book("Gyan Ganga"|"Way of Living"|""), language_mentioned, address.
Return a JSON object with "rows":[{{id,country,phone,book,language_mentioned,address,translated,lang,confidence}}].
- Set "translated" to the input text unchanged.
- Set "lang" to "en" and "confidence" to 1 if confident; else lower confidence in [0,1].

Rules: If unknown, use "" (empty). Never add extra keys.

Rows:
{lines}
""".strip()

    output_rows = _llm_chat_jsonl(prompt, temperature=0, max_tokens=2800)
    out_map = {int(o.get("id", -1)): o for o in output_rows if "id" in o}
    aligned: List[dict] = []
    for i in range(len(texts_en)):
        o = out_map.get(i) or {}
        aligned.append({
            "translated": str(o.get("translated", texts_en[i] or "")),
            "lang": str(o.get("lang", "en")),
            "confidence": float(o.get("confidence", 1.0) or 1.0),
            "country": str(o.get("country", "")),
            "phone": str(o.get("phone", "")),
            "book": str(o.get("book", "")),
            "language_mentioned": str(o.get("language_mentioned", "")),
            "address": str(o.get("address", "")),
        })
    return aligned

# ============================= Unified translation dispatcher =============================
def translate_and_detect(texts: List[str], to_lang: str = "en", provider: Optional[str] = None) -> List[dict]:
    prov = (provider or TRANSLATE_PROVIDER or "llm").lower()

    if prov == "ms":
        trans = _translate_and_detect_ms(texts, to_lang=to_lang)
        translated = [r.get("translated", "") for r in trans]
        ents_full: List[dict] = []
        for i in range(0, len(translated), LLM_MAX_BATCH):
            part = translated[i:i + LLM_MAX_BATCH]
            try:
                ents = llm_entities_only_batch(part)
            except Exception as e:
                log.warning({"event":"llm-entities-failed", "slice": [i, i+len(part)], "error": str(e)})
                ents = [{
                    "translated": p,
                    "lang": "en",
                    "confidence": 0.0,
                    "country":"", "phone":"", "book":"", "language_mentioned":"", "address":""
                } for p in part]
            ents_full.extend(ents)
            del part; gc.collect()

        out: List[dict] = []
        for r, e in zip(trans, ents_full):
            out.append({
                "translated": e.get("translated", r.get("translated","")),
                "lang": r.get("lang", ""),
                "confidence": float(r.get("confidence", 0.0) or 0.0),
                "country": e.get("country",""),
                "phone": e.get("phone",""),
                "book": e.get("book",""),
                "language_mentioned": e.get("language_mentioned",""),
                "address": e.get("address",""),
            })
        return out

    out_full: List[dict] = []
    for i in range(0, len(texts), LLM_MAX_BATCH):
        batch = texts[i:i + LLM_MAX_BATCH]
        log.info({"event":"llm-chunk", "start": i, "end": i+len(batch), "size": len(batch)})
        res = llm_translate_and_extract_batch(batch, to_lang=to_lang)
        out_full.extend(res)
        del batch, res; gc.collect()
    return out_full

# ============================= Text column choice =============================
TEXT_COL_CANDIDATES = {"text", "message", "content", "description", "body"}

def choose_text_col_from_header(headers: List[str], requested: Optional[str]) -> int:
    if requested:
        for idx, h in enumerate(headers):
            if h == requested or h.strip().lower() == requested.strip().lower():
                return idx
    lowered = [h.strip().lower() for h in headers]
    for i, h in enumerate(lowered):
        if h in TEXT_COL_CANDIDATES:
            return i
    return 0 if headers else 0

# ============================= Streaming readers =============================
def iter_csv_rows(path: str, text_column: Optional[str]) -> Iterable[List[str]]:
    first = True
    for chunk in pd.read_csv(
        path,
        chunksize=CSV_CHUNK,
        dtype=str,
        encoding="utf-8",
        on_bad_lines="skip",
        engine="python",
    ):
        for c in chunk.columns:
            chunk[c] = chunk[c].astype(str).fillna("")
        if first:
            headers = [str(c) if c is not None else "" for c in chunk.columns.tolist()]
            yield headers
            first = False
        for _, row in chunk.iterrows():
            yield [row.get(c, "") for c in chunk.columns]
        del chunk; gc.collect()

def iter_excel_rows(path: str, text_column: Optional[str]) -> Iterable[List[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            yield ["text"]
            for r in rows:
                vals = [(v if v is not None else "") for v in (r or [])]
                text = " ".join([str(x) for x in vals if str(x).strip() != ""])
                yield [text]
            return
        headers = [str(h) if h is not None else "" for h in headers]
        yield headers
        for r in rows:
            vals = [(v if v is not None else "") for v in (r or [])]
            if len(vals) < len(headers):
                vals = list(vals) + [""] * (len(headers) - len(vals))
            else:
                vals = list(vals[:len(headers)])
            vals = [str(v) if v is not None else "" for v in vals]
            yield vals
    finally:
        wb.close()

# ============================= Row estimation =============================
def estimate_total_rows(in_path: str, original_name: str) -> Optional[int]:
    name = (original_name or "").lower()
    try:
        if name.endswith(".csv"):
            with open(in_path, "rb") as fh:
                data = fh.read()
            total = data.count(b"\n")
            return max(0, total - 1)
        if name.endswith((".xlsx", ".xlsm", ".xls")):
            wb = load_workbook(in_path, read_only=True, data_only=True)
            try:
                ws = wb.active
                mr = int(ws.max_row or 0)
                return max(0, mr - 1)
            finally:
                wb.close()
    except Exception as e:
        log.warning({"event":"estimate-total-rows-failed","file": original_name, "error": str(e)})
    return None

# ============================= Core streaming processor =============================
def process_rows_streaming(
    rows_iter: Iterable[List[str]],
    requested_text_col: Optional[str],
    to_lang: str,
    provider: Optional[str],
    outfile: io.TextIOBase,
    on_progress: Optional[Callable[[int, int], None]] = None,  # rows_added, batches_added
):
    writer = csv.writer(outfile, lineterminator="\n")
    header = next(rows_iter, None)
    if header is None:
        writer.writerow(["text","translated_en","source_lang","translation_confidence","was_translated",
                         "translation_needs_review","country","phone","book","language_mentioned","address","dialing_code"])
        return

    header = [str(h) if h is not None else "" for h in header]
    writer.writerow(header + ["translated_en","source_lang","translation_confidence","was_translated",
                              "translation_needs_review","country","phone","book","language_mentioned","address","dialing_code"])

    text_col_idx = choose_text_col_from_header(header, requested_text_col)

    buf_rows: List[List[str]] = []
    buf_texts: List[str] = []

    def flush_batch():
        if not buf_rows:
            return
        results = translate_and_detect(buf_texts, to_lang=to_lang, provider=provider)
        written = 0
        for original_row, res in zip(buf_rows, results):
            t_en = res.get("translated","")
            lang = (res.get("lang","") or "").lower()
            conf = float(res.get("confidence", 0.0) or 0.0)
            was_trans = bool(lang and lang != "en")
            needs_review = bool(was_trans and conf < ENTITY_CONF_THRESHOLD)

            country = res.get("country","")
            phone = res.get("phone","")
            book = res.get("book","")
            language_mentioned = res.get("language_mentioned","")
            address = res.get("address","")
            dial = country_to_dial(country, phone)

            writer.writerow(original_row + [
                t_en, lang, f"{conf:.3f}", "true" if was_trans else "false",
                "true" if needs_review else "false",
                country, phone, book, language_mentioned, address, dial
            ])
            written += 1
        if on_progress:
            on_progress(written, 1)
        buf_rows.clear(); buf_texts.clear()
        gc.collect()

    for row in rows_iter:
        if len(row) < len(header):
            row = list(row) + [""] * (len(header) - len(row))
        else:
            row = list(row[:len(header)])
        text = str(row[text_col_idx] or "").strip()
        buf_rows.append(row)
        buf_texts.append(text)
        if len(buf_rows) >= LLM_MAX_BATCH:
            flush_batch()

    flush_batch()

# ============================= File helpers =============================
def stream_file_iter(path: str, chunk_size: int = 1024*256):
    with open(path, "rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            yield chunk

def guess_ext(name: str) -> str:
    nl = (name or "").lower()
    if nl.endswith(".csv"): return ".csv"
    if nl.endswith(".xlsx"): return ".xlsx"
    if nl.endswith(".xlsm"): return ".xlsm"
    if nl.endswith(".xls"): return ".xls"
    return os.path.splitext(nl)[1] or ".csv"

def process_local_file_to_csv(
    in_path: str,
    original_name: str,
    text_column: Optional[str],
    provider: Optional[str],
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> str:
    ext = guess_ext(original_name)
    out_fd, out_path = tempfile.mkstemp(prefix="enriched_", suffix=".csv")
    os.close(out_fd)
    with open(out_path, "w", encoding="utf-8", newline="") as out_io:
        rows = iter_csv_rows(in_path, text_column) if ext == ".csv" else iter_excel_rows(in_path, text_column)
        process_rows_streaming(
            rows_iter=rows,
            requested_text_col=text_column,
            to_lang="en",
            provider=provider,
            outfile=out_io,
            on_progress=on_progress,
        )
    return out_path

def upload_processed_blob(blob_service: BlobServiceClient, processed_path: str, processed_filename: str) -> None:
    ensure_container(blob_service, "processed")
    client = blob_service.get_blob_client(container="processed", blob=processed_filename)
    with open(processed_path, "rb") as fh:
        client.upload_blob(fh, overwrite=True, content_settings=ContentSettings(content_type="text/csv"))
    log.info({"event":"blob-upload", "blob": f"processed/{processed_filename}", "bytes": os.path.getsize(processed_path)})

# ============================= Async Job model =============================
class JobState(str, Enum):
    queued = "queued"
    running = "running"
    done = "done"
    error = "error"

@dataclass
class Job:
    id: str
    filename: str
    text_column: Optional[str]
    provider: Optional[str]
    started_at: float = field(default_factory=time.time)
    finished_at: Optional[float] = None
    state: JobState = JobState.queued
    error: Optional[str] = None
    incoming_blob: Optional[str] = None
    processed_filename: Optional[str] = None
    processed_path: Optional[str] = None
    rows_processed: int = 0
    batches_processed: int = 0
    total_rows: Optional[int] = None
    run_id: str = field(default_factory=lambda: uuid.uuid4().hex[:8])

JOBS: Dict[str, Job] = {}

async def _run_upload_job(job: Job, data: bytes):
    job.state = JobState.running
    req_id = job.run_id
    in_path = None
    try:
        safe_base = re.sub(r"[^A-Za-z0-9_.-]", "_", os.path.basename(job.filename or "upload.csv"))
        ts = time.strftime("%Y%m%d-%H%M%S")
        incoming_name = f"{ts}_{safe_base}"

        in_fd, in_path = tempfile.mkstemp(prefix="upload_", suffix=os.path.splitext(safe_base)[1] or ".csv")
        os.close(in_fd)
        await anyio.to_thread.run_sync(lambda: open(in_path, "wb").write(data))

        job.total_rows = estimate_total_rows(in_path, safe_base)
        json_log("processing_started", run_id=req_id, file=safe_base, total_rows=job.total_rows)

        blob_service = get_blob_client()
        if blob_service is not None:
            try:
                def _upload_incoming():
                    ensure_container(blob_service, "incoming")
                    ctype = "text/csv" if safe_base.lower().endswith(".csv") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    in_client = blob_service.get_blob_client(container="incoming", blob=incoming_name)
                    with open(in_path, "rb") as fh:
                        in_client.upload_blob(fh, overwrite=True, content_settings=ContentSettings(content_type=ctype))
                await anyio.to_thread.run_sync(_upload_incoming)
                log.info({"event":"upload-to-incoming", "blob": f"incoming/{incoming_name}", "bytes": os.path.getsize(in_path), "req_id": req_id})
                job.incoming_blob = f"incoming/{incoming_name}"
            except Exception as e:
                log.warning({"event":"blob-route-failed", "error": str(e), "req_id": req_id})

        def _on_progress(rows_added: int, batches_added: int):
            job.rows_processed += int(rows_added)
            job.batches_processed += int(batches_added)
            json_log(
                "progress",
                run_id=req_id,
                rows_processed=job.rows_processed,
                batches_processed=job.batches_processed,
                total_rows=job.total_rows,
                message=f"Processed +{rows_added} rows (flush)."
            )

        t0 = time.time()
        out_path = await anyio.to_thread.run_sync(
            lambda: process_local_file_to_csv(
                in_path,
                original_name=safe_base,
                text_column=job.text_column,
                provider=job.provider,
                on_progress=_on_progress,
            )
        )
        ms = int((time.time() - t0) * 1000)
        processed_filename = os.path.splitext(safe_base)[0] + "_enriched.csv"
        log.info({"event":"process-upload", "file": safe_base, "processed": processed_filename, "ms": ms, "provider": job.provider or TRANSLATE_PROVIDER, "req_id": req_id})

        if blob_service is not None:
            try:
                await anyio.to_thread.run_sync(lambda: upload_processed_blob(blob_service, out_path, processed_filename))
            except Exception as e:
                log.warning({"event":"processed-upload-skip", "error": str(e), "req_id": req_id})

        job.processed_filename = processed_filename
        job.processed_path = out_path
        job.state = JobState.done
        job.finished_at = time.time()
        json_log("processing_completed", run_id=req_id, rows_processed=job.rows_processed, total_rows=job.total_rows)
    except Exception as e:
        log.exception({"event":"async-job-failed", "error": str(e), "req_id": req_id})
        job.error = str(e)
        job.state = JobState.error
        job.finished_at = time.time()
        json_log("processing_failed", run_id=req_id, error=str(e), rows_processed=job.rows_processed, total_rows=job.total_rows)
    finally:
        try:
            if in_path and os.path.exists(in_path):
                os.remove(in_path)
        except Exception:
            pass

# ============================= Endpoints =============================
@app.get("/health")
async def health_check():
    return {"status": "healthy"}

@app.get("/debug-env")
async def debug_env():
    return {
        "AZURE_TRANSLATOR_ENDPOINT": os.environ.get("AZURE_TRANSLATOR_ENDPOINT"),
        "AZURE_TRANSLATOR_KEY": "REDACTED" if os.environ.get("AZURE_TRANSLATOR_KEY") else "MISSING",
        "AZURE_TRANSLATOR_REGION": os.environ.get("AZURE_TRANSLATOR_REGION"),
        "AZURE_OPENAI_ENDPOINT": os.environ.get("AZURE_OPENAI_ENDPOINT"),
        "AZURE_OPENAI_API_KEY": "REDACTED" if os.environ.get("AZURE_OPENAI_API_KEY") else "MISSING",
        "AZURE_OPENAI_DEPLOYMENT": AOAI_DEP,
        "AZURE_OPENAI_API_VERSION": AOAI_VER,
        "TRANSLATE_PROVIDER": TRANSLATE_PROVIDER,
        "LLM_MAX_BATCH": LLM_MAX_BATCH,
        "CSV_CHUNK": CSV_CHUNK,
        "STORAGE_ACCOUNT_NAME": os.environ.get("STORAGE_ACCOUNT_NAME"),
        "AZURE_STORAGE_CONNECTION_STRING": "REDACTED" if os.environ.get("AZURE_STORAGE_CONNECTION_STRING") else "MISSING",
        "CORS_ALLOW_ORIGINS": allow_origins,
    }

@app.get("/logs")
async def logs_api(limit: int = Query(500, ge=1, le=2000)):
    buf = ring_handler.buf[-limit:]
    return {"items": buf, "count": len(buf)}

@app.post("/translate")
async def translate_api(payload: dict):
    req_id = uuid.uuid4().hex[:8]
    texts = payload.get("texts") or []
    to = payload.get("to", "en")
    provider = (payload.get("provider") or "").lower() or None
    if not isinstance(texts, list) or len(texts) == 0:
        raise HTTPException(400, "Provide texts: []")
    t0 = time.time()
    out = await anyio.to_thread.run_sync(lambda: translate_and_detect([str(x) for x in texts], to_lang=to, provider=provider))
    log.info({"event": "translate", "n": len(texts), "ms": int((time.time() - t0) * 1000), "provider": provider or TRANSLATE_PROVIDER, "req_id": req_id})
    return {"rows": out}

# ---------- Synchronous (kept for compatibility; can timeout for big files) ----------
@app.post("/process-xlsx")
async def process_xlsx(
    blob_name: str = Query(..., description="e.g., 'incoming/sample.xlsx' or 'incoming/sample.csv'"),
    text_column: Optional[str] = Query(None),
    provider: Optional[str] = Query(None)
):
    req_id = uuid.uuid4().hex[:8]
    if not blob_name.lower().endswith((".xlsx", ".xlsm", ".xls", ".csv")) or not blob_name.startswith("incoming/"):
        raise HTTPException(400, "Provide valid blob_name (e.g., 'incoming/sample.xlsx' or 'incoming/sample.csv')")

    blob_service = get_blob_client()
    if blob_service is None:
        raise HTTPException(500, "Storage not configured")

    ensure_container(blob_service, "incoming")
    ensure_container(blob_service, "processed")

    in_tmp_fd, in_tmp_path = tempfile.mkstemp(prefix="incoming_", suffix=os.path.splitext(blob_name)[1])
    os.close(in_tmp_fd)
    try:
        def _download():
            in_client = blob_service.get_blob_client(container="incoming", blob=blob_name.replace("incoming/", ""))
            log.info({"event": "blob-download-start", "blob": blob_name, "req_id": req_id})
            with open(in_tmp_path, "wb") as fh:
                downloader = in_client.download_blob()
                downloader.readinto(fh)
            log.info({"event": "blob-download-complete", "blob": blob_name, "bytes": os.path.getsize(in_tmp_path), "req_id": req_id})

        await anyio.to_thread.run_sync(_download)

        total_rows = estimate_total_rows(in_tmp_path, blob_name)
        json_log("processing_started", run_id=req_id, file=blob_name, total_rows=total_rows)

        rows_processed = 0
        batches_processed = 0
        def _on_progress(rows_added: int, batches_added: int):
            nonlocal rows_processed, batches_processed
            rows_processed += rows_added
            batches_processed += batches_added
            json_log("progress", run_id=req_id, rows_processed=rows_processed, batches_processed=batches_processed, total_rows=total_rows)

        t0 = time.time()
        out_path = await anyio.to_thread.run_sync(
            lambda: process_local_file_to_csv(in_tmp_path, original_name=blob_name, text_column=text_column, provider=provider, on_progress=_on_progress)
        )
        ms = int((time.time() - t0) * 1000)
        processed_filename = os.path.basename(blob_name).rsplit(".", 1)[0] + "_enriched.csv"
        log.info({"event": "process-xlsx", "blob": blob_name, "processed": processed_filename, "ms": ms, "provider": provider or TRANSLATE_PROVIDER, "req_id": req_id})

        await anyio.to_thread.run_sync(lambda: upload_processed_blob(blob_service, out_path, processed_filename))

        json_log("processing_completed", run_id=req_id, rows_processed=rows_processed, total_rows=total_rows)

        headers = {"Content-Disposition": f'attachment; filename="{processed_filename}"'}
        return StreamingResponse(stream_file_iter(out_path), media_type="text/csv", headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        log.exception({"event": "process-failed-soft", "blob": blob_name, "error": str(e), "req_id": req_id})
        json_log("processing_failed", run_id=req_id, error=str(e))
        return PlainTextResponse(f"error,{str(e)}\n", media_type="text/csv", status_code=200)
    finally:
        try:
            os.remove(in_tmp_path)
        except Exception:
            pass

@app.post("/process-upload")
async def process_upload(
    file: UploadFile = File(...),
    text_column: Optional[str] = Form(None),
    provider: Optional[str] = Form(None)
):
    req_id = uuid.uuid4().hex[:8]
    original = file.filename or "upload.csv"
    safe_base = re.sub(r"[^A-Za-z0-9_.-]", "_", os.path.basename(original))
    ts = time.strftime("%Y%m%d-%H%M%S")
    incoming_name = f"{ts}_{safe_base}"

    in_fd, in_path = tempfile.mkstemp(prefix="upload_", suffix=os.path.splitext(safe_base)[1] or ".csv")
    os.close(in_fd)
    try:
        data = await file.read()
        await anyio.to_thread.run_sync(lambda: open(in_path, "wb").write(data))
        del data; gc.collect()

        total_rows = estimate_total_rows(in_path, safe_base)
        json_log("processing_started", run_id=req_id, file=safe_base, total_rows=total_rows)
        rows_processed = 0
        batches_processed = 0
        def _on_progress(rows_added: int, batches_added: int):
            nonlocal rows_processed, batches_processed
            rows_processed += rows_added
            batches_processed += batches_added
            json_log("progress", run_id=req_id, rows_processed=rows_processed, batches_processed=batches_processed, total_rows=total_rows)

        blob_service = get_blob_client()

        if blob_service is not None:
            try:
                def _upload_incoming():
                    ensure_container(blob_service, "incoming")
                    ctype = "text/csv" if safe_base.lower().endswith(".csv") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    in_client = blob_service.get_blob_client(container="incoming", blob=incoming_name)
                    with open(in_path, "rb") as fh:
                        in_client.upload_blob(fh, overwrite=True, content_settings=ContentSettings(content_type=ctype))
                await anyio.to_thread.run_sync(_upload_incoming)
                log.info({"event": "upload-to-incoming", "blob": f"incoming/{incoming_name}", "bytes": os.path.getsize(in_path), "req_id": req_id})
            except Exception as e:
                log.warning({"event": "blob-route-failed", "error": str(e), "req_id": req_id})

        t0 = time.time()
        out_path = await anyio.to_thread.run_sync(lambda: process_local_file_to_csv(in_path, original_name=safe_base, text_column=text_column, provider=provider, on_progress=_on_progress))
        ms = int((time.time() - t0) * 1000)
        processed_filename = os.path.splitext(safe_base)[0] + "_enriched.csv"
        log.info({"event":"process-upload", "file": safe_base, "processed": processed_filename, "ms": ms, "provider": provider or TRANSLATE_PROVIDER, "req_id": req_id})

        if blob_service is not None:
            try:
                await anyio.to_thread.run_sync(lambda: upload_processed_blob(blob_service, out_path, processed_filename))
            except Exception as e:
                log.warning({"event":"processed-upload-skip", "error": str(e), "req_id": req_id})

        json_log("processing_completed", run_id=req_id, rows_processed=rows_processed, total_rows=total_rows)

        headers = {"Content-Disposition": f'attachment; filename="{processed_filename}"'}
        return StreamingResponse(stream_file_iter(out_path), media_type="text/csv", headers=headers)

    except HTTPException:
        raise
    except Exception as e:
        log.exception({"event":"process-upload-failed-soft", "error": str(e), "req_id": req_id})
        json_log("processing_failed", run_id=req_id, error=str(e))
        return PlainTextResponse(f"error,{str(e)}\n", media_type="text/csv", status_code=200)
    finally:
        try: os.remove(in_path)
        except Exception: pass

# ---------- Async (recommended) – avoids front-door timeout ----------
@app.post("/process-upload-async")
async def process_upload_async(
    file: UploadFile = File(...),
    text_column: Optional[str] = Form(None),
    provider: Optional[str] = Form(None)
):
    data = await file.read()
    job_id = uuid.uuid4().hex
    job = Job(id=job_id, filename=file.filename or "upload.csv", text_column=text_column, provider=provider)
    JOBS[job_id] = job
    asyncio.create_task(_run_upload_job(job, data))
    return {"job_id": job_id, "state": job.state}

@app.get("/jobs/{job_id}")
async def job_status(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "job not found")
    return {
        "job_id": job.id,
        "state": job.state,
        "filename": job.filename,
        "incoming_blob": job.incoming_blob,
        "processed_filename": job.processed_filename,
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "error": job.error,
        "rows_processed": job.rows_processed,
        "batches_processed": job.batches_processed,
        "total_rows": job.total_rows,
        "run_id": job.run_id,
    }

@app.get("/jobs/{job_id}/progress")
async def job_progress(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "job not found")
    return {
        "job_id": job.id,
        "state": job.state,
        "rows_processed": job.rows_processed,
        "batches_processed": job.batches_processed,
        "total_rows": job.total_rows,
        "run_id": job.run_id,
    }

@app.get("/jobs/{job_id}/download")
async def job_download(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "job not found")
    if job.state != JobState.done or not job.processed_path:
        raise HTTPException(409, "not ready")
    fname = job.processed_filename or "output.csv"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"'}
    return StreamingResponse(stream_file_iter(job.processed_path), media_type="text/csv", headers=headers)
